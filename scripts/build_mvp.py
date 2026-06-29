from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
SEED_PATH = ROOT / "scripts" / "seed_companies.json"
DATA_DIR = ROOT / "data"
PUBLIC_DATA_DIR = ROOT / "public" / "data"

HEADERS = [
    ("company_id", "Įrašo ID"),
    ("company_code", "Įmonės kodas"),
    ("name", "Įmonės pavadinimas"),
    ("brand", "Prekės ženklas"),
    ("market_role", "Rinkos vaidmuo"),
    ("city", "Miestas"),
    ("municipality", "Savivaldybė"),
    ("address", "Adresas"),
    ("latitude", "Platuma"),
    ("longitude", "Ilguma"),
    ("activity_type", "Veiklos tipas"),
    ("activity_label", "Veikla"),
    ("concrete_plant_name", "Betono mazgo / gamyklos pavadinimas"),
    ("concrete_plant_capacity", "Betono mazgo našumas"),
    ("concrete_plant_description", "Viešas betono mazgo / gamyklos aprašymas"),
    ("classification_confidence", "Klasifikavimo pasitikėjimas"),
    ("revenue_latest", "Įmonės apyvarta"),
    ("revenue_year", "Apyvartos metai"),
    ("employees_latest", "Darbuotojų skaičius"),
    ("vehicles_latest", "Automobilių / transporto priemonių skaičius"),
    ("website", "Interneto svetainė"),
    ("source_url", "Pagrindinis šaltinis"),
    ("revenue_source_url", "Apyvartos šaltinis"),
    ("employees_source_url", "Darbuotojų šaltinis"),
    ("vehicles_source_url", "Transporto šaltinis"),
    ("plant_source_url", "Betono mazgo / gamyklos šaltinis"),
    ("manual_note", "Pastaba / rankinis papildymas"),
    ("last_updated", "Paskutinio atnaujinimo data"),
]

KEYS = [key for key, _label in HEADERS]
LABELS = [label for _key, label in HEADERS]


def load_companies() -> list[dict]:
    with SEED_PATH.open(encoding="utf-8") as file:
        return json.load(file)


def write_json(companies: list[dict]) -> None:
    PUBLIC_DATA_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "scope": {
            "country": "Lietuva",
            "cities": "Visa Lietuva",
            "excluded_keywords": ["trinkelės"],
        },
        "companies": companies,
    }
    (PUBLIC_DATA_DIR / "companies.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def style_header(row) -> None:
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(color="FFFFFF", bold=True)
    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")


def auto_width(ws) -> None:
    for column in ws.columns:
        max_len = 0
        letter = column[0].column_letter
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 42)


def write_workbook(companies: list[dict]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Įmonės"

    ws.append(LABELS)
    for company in companies:
        ws.append([company.get(key) for key in KEYS])

    style_header(ws[1])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    table = Table(displayName="ImonesTable", ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    thin = Side(style="thin", color="E5E7EB")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    for row in ws.iter_rows(min_row=2, min_col=9, max_col=10):
        for cell in row:
            cell.number_format = "0.000000"
    for row in ws.iter_rows(min_row=2, min_col=16, max_col=16):
        for cell in row:
            cell.number_format = "0.00"
    for row in ws.iter_rows(min_row=2, min_col=17, max_col=20):
        for cell in row:
            cell.number_format = "#,##0"

    auto_width(ws)

    meta = wb.create_sheet("Metaduomenys")
    metadata_rows = [
        ["Projektas", "Konkurentų žemėlapis"],
        ["Apimtis", "Visa Lietuva"],
        ["Įtraukta", "Betono mišiniai, betono mazgai, gelžbetonis ir surenkamas gelžbetonis"],
        ["Neįtraukta", "Trinkelės, jei tai pagrindinė veikla"],
        ["Duomenų modelis", "Excel yra MVP duomenų failas; JSON generuojamas žemėlapiui"],
        ["Sugeneruota", datetime.now(timezone.utc).isoformat()],
    ]
    for row in metadata_rows:
        meta.append(row)
    for cell in meta["A"]:
        cell.fill = PatternFill("solid", fgColor="0F766E")
        cell.font = Font(color="FFFFFF", bold=True)
    auto_width(meta)

    log = wb.create_sheet("Importo žurnalas")
    log.append(["Paleidimo laikas", "Būsena", "Šaltinis", "Eilučių skaičius", "Pastabos"])
    log.append(
        [
            datetime.now(timezone.utc).isoformat(),
            "gerai",
            "scripts/seed_companies.json",
            len(companies),
            "MVP seed duomenys. Pilnas atvirų duomenų importas prijungiamas kitose iteracijose.",
        ]
    )
    style_header(log[1])
    auto_width(log)

    wb.save(DATA_DIR / "companies.xlsx")


def main() -> None:
    companies = load_companies()
    write_json(companies)
    write_workbook(companies)
    print(f"Built data/companies.xlsx and public/data/companies.json for {len(companies)} rows.")


if __name__ == "__main__":
    main()
